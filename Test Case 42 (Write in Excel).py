import openpyxl

#for same data
file="D:/OneDrive/Excel/Data.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet1"]
for r in range(1,6):
    for c in range(1,4):
        sheet.cell(r,c).value="Welcome"

workbook.save(file)

#for different data
file="D:/OneDrive/Excel/Data.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet2"]
sheet.cell(1,1).value=123
sheet.cell(1,2).value="Usha"
sheet.cell(2,1).value=456
sheet.cell(2,2).value="Anaye"

workbook.save(file)