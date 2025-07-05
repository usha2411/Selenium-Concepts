#file--workbook--sheets--rows--cells

import openpyxl
file="D:/OneDrive/Excel/Project1.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet1"]
rows=sheet.max_row
columns=sheet.max_column

#read  all row and col
for r in range(1,rows+1):
    for c in range(1,columns+1):
        print(sheet.cell(r,c).value,end="   ")
    print()
